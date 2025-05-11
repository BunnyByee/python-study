import requests
from bs4 import BeautifulSoup
import openpyxl

response = requests.get("https://finance.naver.com/marketindex/?tabSel=exchange#tab_section")
html = response.text
soup = BeautifulSoup(html,'html.parser')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'codingon_ex'
ws['A1'] = '통화'
ws['B1'] = '환율'

# m_title = soup.select('.market1 .data .h_lst .blind')
# m_num = soup.select('.market1 .data .value')

# for i in m_title :
#     txt = i.text
#     clear_text = ' '.join(txt.split()[1:])
#     print(clear_text)

# for j in m_num :
#     print(j.text)

# for i, j in zip(m_title, m_num) :
#     txt = i.text
#     num = j.text
#     clear_txt = ' '.join(txt.split()[1:])
#     print(clear_txt,num)

all_li = soup.select('.market1 ul.data_lst > li')
# print(all_li)

for li in all_li:
    title = li.select_one('span.blind')
    value = li.select_one('span.value')
    print(title.string.split()[-1], value.text)
    ws.append([title.string.split()[-1], value.text])

wb.save('codingon_ex.xlsx')



