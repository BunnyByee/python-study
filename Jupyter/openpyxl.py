# %%
# %pip install openpyxl

# %%
import openpyxl

# %%
wb = openpyxl.Workbook()
ws = wb.create_sheet('codingon')

ws['A1'] = '이름'
ws['B1'] = '영어이름'

ws['A2'] = '홍길동'
ws['B2'] = 'Hong'

wb.save('codingon.xlsx')

# %%
wb = openpyxl.load_workbook('codingon.xlsx')

ws = wb['codingon']

ws['A3'] = '1111'
ws['B3'] = '34324'

wb.save('codingon.xlsx')

# %%
wb = openpyxl.Workbook()

ws = wb.active
ws.title = 'Data'

data = [['이름', 'name'],['가나다', 'ganada'],['밥','Bob']]

for row in data :
    ws.append(row)

wb.save('append_ex.xlsx')

# %%
import pandas as pd
a = pd.read_excel('append_ex.xlsx')
print(a)


